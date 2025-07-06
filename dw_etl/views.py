# dw_etl/views.py
from django.shortcuts import render
from django.db.models import F, Case, When, Value, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse # Importar JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import json # Importar la librería json

from .models import HechosEconomicos, DimPais, DimIndicadorEconomico, DimFecha

def dashboard_view(request):
    """
    Vista para mostrar un dashboard con los datos económicos históricos.
    Permite seleccionar un año para visualizar y muestra un gráfico de evolución.
    """
    # Recuperar los objetos de indicadores económicos
    try:
        inflacion_indicador = DimIndicadorEconomico.objects.get(nombre_indicador='Inflación')
        pib_indicador = DimIndicadorEconomico.objects.get(nombre_indicador='Crecimiento PIB')
        tipo_cambio_indicador = DimIndicadorEconomico.objects.get(nombre_indicador='Tipo de Cambio Dólar')
        ipc_indicador = DimIndicadorEconomico.objects.get(nombre_indicador='IPC')
    except DimIndicadorEconomico.DoesNotExist:
        return render(request, 'dw_etl/dashboard.html', {'error_message': 'Uno o más indicadores económicos no se encontraron en la base de datos. Por favor, asegúrate de que el comando populate_dw se ejecutó correctamente.'})

    # Obtener todos los años disponibles en DimFecha para el selector
    available_years_with_data = HechosEconomicos.objects.values_list('id_fecha__anio', flat=True).distinct().order_by('-id_fecha__anio')
    available_years = sorted(list(available_years_with_data), reverse=True)

    # Determinar el año a mostrar (por defecto, el más reciente disponible)
    selected_year = request.GET.get('year')
    if selected_year:
        try:
            selected_year = int(selected_year)
            # Si el año seleccionado no está en los disponibles, usar el más reciente
            if selected_year not in available_years:
                selected_year = available_years[0] if available_years else datetime.now().year
        except ValueError:
            # Si hay un error en el valor del año, usar el más reciente
            selected_year = available_years[0] if available_years else datetime.now().year
    else:
        # Si no se selecciona año, usar el más reciente disponible
        selected_year = available_years[0] if available_years else datetime.now().year
    
    # Obtener el objeto DimFecha para el año seleccionado
    try:
        fecha_seleccionada = DimFecha.objects.get(anio=selected_year)
    except DimFecha.DoesNotExist:
        # Esto no debería ocurrir si available_years se basa en hechos existentes
        return render(request, 'dw_etl/dashboard.html', {'error_message': f'No se encontraron datos de fecha para el año {selected_year}. Por favor, selecciona otro año o ejecuta populate_dw.'})


    # Preparar un diccionario para almacenar los datos por país para el año seleccionado
    dashboard_data_dict = {}
    for pais in DimPais.objects.all().order_by('nombre_pais'):
        dashboard_data_dict[pais.nombre_pais] = {
            'nombre_pais': pais.nombre_pais, 'iso': pais.codigo_iso,
            'inflacion': None, 'pib_crecimiento': None, 'tipo_cambio': None, 'ipc': None,
        }

    # Recuperar todos los hechos económicos para el año seleccionado
    hechos = HechosEconomicos.objects.filter(id_fecha=fecha_seleccionada).select_related('id_pais', 'id_indicador')

    for hecho in hechos:
        pais_nombre = hecho.id_pais.nombre_pais
        indicador_nombre = hecho.id_indicador.nombre_indicador

        if pais_nombre in dashboard_data_dict:
            if indicador_nombre == 'Inflación':
                dashboard_data_dict[pais_nombre]['inflacion'] = hecho.porcentaje_inflacion
            elif indicador_nombre == 'Crecimiento PIB':
                dashboard_data_dict[pais_nombre]['pib_crecimiento'] = hecho.variacion_pib_anual
            elif indicador_nombre == 'Tipo de Cambio Dólar':
                dashboard_data_dict[pais_nombre]['tipo_cambio'] = hecho.tipo_cambio_usd_local_promedio_cierre
            elif indicador_nombre == 'IPC':
                dashboard_data_dict[pais_nombre]['ipc'] = hecho.ipc_o_devaluacion

    # Convertir el diccionario a una lista para pasar al template
    dashboard_data = list(dashboard_data_dict.values())

    # --- Lógica de Análisis Comparativo (para el año seleccionado) ---
    best_inflation_country = None
    min_inflation = float('inf')

    for data in dashboard_data:
        if data['inflacion'] is not None and data['inflacion'] != 'N/D':
            try:
                current_inflation = float(data['inflacion'])
                if current_inflation < min_inflation:
                    min_inflation = current_inflation
                    best_inflation_country = data['nombre_pais']
            except ValueError:
                pass # Ignorar si el valor no es un número válido

    # --- Preparar datos para el gráfico de evolución histórica (Chile) ---
    chile_data = {}
    try:
        chile_pais = DimPais.objects.get(nombre_pais='Chile')
        chile_inflation_facts = HechosEconomicos.objects.filter(
            id_pais=chile_pais,
            id_indicador=inflacion_indicador,
            porcentaje_inflacion__isnull=False
        ).select_related('id_fecha').order_by('id_fecha__anio')

        chile_pib_facts = HechosEconomicos.objects.filter(
            id_pais=chile_pais,
            id_indicador=pib_indicador,
            variacion_pib_anual__isnull=False
        ).select_related('id_fecha').order_by('id_fecha__anio')

        chile_inflation_history = [{'year': f.id_fecha.anio, 'value': float(f.porcentaje_inflacion)} for f in chile_inflation_facts]
        chile_pib_history = [{'year': f.id_fecha.anio, 'value': float(f.variacion_pib_anual)} for f in chile_pib_facts]

        # Combinar y asegurar que los años estén alineados para el gráfico
        all_chile_years = sorted(list(set([d['year'] for d in chile_inflation_history] + [d['year'] for d in chile_pib_history])))
        
        chile_chart_data = {
            'labels': all_chile_years,
            'inflation_data': [next((d['value'] for d in chile_inflation_history if d['year'] == year), None) for year in all_chile_years],
            'pib_data': [next((d['value'] for d in chile_pib_history if d['year'] == year), None) for year in all_chile_years],
        }

    except DimPais.DoesNotExist:
        chile_chart_data = None
        # self.stdout.write(self.style.WARNING('No se encontró el país Chile para el gráfico histórico.'))
    except DimIndicadorEconomico.DoesNotExist:
        chile_chart_data = None
        # self.stdout.write(self.style.WARNING('No se encontraron indicadores para el gráfico histórico de Chile.'))
    except Exception as e:
        chile_chart_data = None
        # self.stdout.write(self.style.ERROR(f'Error al preparar datos históricos de Chile: {e}'))


    # Preparar el contexto para el template
    context = {
        'dashboard_data': dashboard_data,
        'current_year': selected_year, # Ahora es el año seleccionado
        'available_years': available_years, # Años para el selector
        'best_inflation_country': best_inflation_country,
        'min_inflation_value': min_inflation if min_inflation != float('inf') else 'N/D',
        'chile_chart_data': chile_chart_data, # Datos para el gráfico de Chile
        'analysis_message_devaluation_recession': "La identificación de periodos donde coinciden una fuerte devaluación con recesión requiere un análisis de series de tiempo y la definición de umbrales para 'fuerte devaluación' y 'recesión' (ej. dos trimestres consecutivos de crecimiento negativo del PIB). Este tipo de análisis se realiza de forma más efectiva en herramientas de BI como Power BI, utilizando el archivo Excel exportado que contiene la data histórica completa.",
        'analysis_message_exchange_gdp_relation': "La relación entre el tipo de cambio y el crecimiento del PIB es compleja y varía por país y periodo. Generalmente, una devaluación puede hacer las exportaciones más baratas (impulsando el PIB) pero también encarecer las importaciones (afectando el IPC y el poder adquisitivo). Un análisis de correlación y causalidad a lo largo del tiempo es crucial, y se puede realizar eficientemente en Power BI con los datos históricos exportados.",
        'conclusion_message': "El modelo de Data Warehouse implementado y poblado con datos históricos de la API del Banco Mundial ha demostrado ser efectivo para centralizar y estructurar información económica clave. El dashboard permite una visualización rápida por año y la identificación del país con menor inflación. Para análisis más profundos de tendencias, correlaciones y eventos históricos complejos (como devaluación y recesión, o la relación tipo de cambio-crecimiento), la exportación a Excel para herramientas de BI como Power BI es la vía recomendada, aprovechando la riqueza de la data histórica cargada.",
    }
    return render(request, 'dw_etl/dashboard.html', context)


def export_economic_data_excel(request):
    """
    Vista para exportar TODOS los datos económicos históricos a un archivo Excel.
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="datos_economicos_historicos.xlsx"'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Datos Económicos Históricos"

    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_aligned_text = Alignment(horizontal="center", vertical="center")

    # Encabezados de la tabla
    headers = ["Año", "País", "Código ISO", "Inflación (%)", "Crecimiento PIB (%)", "Tipo Cambio (USD/Local)", "IPC (Índice)"]
    sheet.append(headers)

    # Aplicar estilos a la fila de encabezados
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_aligned_text
        sheet.column_dimensions[get_column_letter(col_num)].width = 20 # Ajustar ancho de columna

    # Obtener TODOS los hechos económicos históricos
    # Usamos select_related para optimizar las consultas a DimFecha, DimPais, DimIndicadorEconomico
    hechos = HechosEconomicos.objects.select_related('id_fecha', 'id_pais', 'id_indicador').order_by('id_fecha__anio', 'id_pais__nombre_pais', 'id_indicador__nombre_indicador')

    # Agrupar los datos para facilitar la exportación por año y país
    export_data_structured = {}
    for hecho in hechos:
        anio = hecho.id_fecha.anio
        pais_nombre = hecho.id_pais.nombre_pais
        indicador_nombre = hecho.id_indicador.nombre_indicador

        if anio not in export_data_structured:
            export_data_structured[anio] = {}
        if pais_nombre not in export_data_structured[anio]:
            export_data_structured[anio][pais_nombre] = {
                'nombre_pais': pais_nombre,
                'iso': hecho.id_pais.codigo_iso,
                'inflacion': None,
                'pib_crecimiento': None,
                'tipo_cambio': None,
                'ipc': None,
            }
        
        if indicador_nombre == 'Inflación':
            export_data_structured[anio][pais_nombre]['inflacion'] = hecho.porcentaje_inflacion
        elif indicador_nombre == 'Crecimiento PIB':
            export_data_structured[anio][pais_nombre]['pib_crecimiento'] = hecho.variacion_pib_anual
        elif indicador_nombre == 'Tipo de Cambio Dólar':
            export_data_structured[anio][pais_nombre]['tipo_cambio'] = hecho.tipo_cambio_usd_local_promedio_cierre
        elif indicador_nombre == 'IPC':
            export_data_structured[anio][pais_nombre]['ipc'] = hecho.ipc_o_devaluacion

    # Escribir los datos en la hoja de cálculo
    for anio in sorted(export_data_structured.keys()):
        for pais_nombre in sorted(export_data_structured[anio].keys()):
            data = export_data_structured[anio][pais_nombre]
            row_data = [
                anio,
                data['nombre_pais'],
                data['iso'],
                data['inflacion'] if data['inflacion'] is not None else 'N/D',
                data['pib_crecimiento'] if data['pib_crecimiento'] is not None else 'N/D',
                data['tipo_cambio'] if data['tipo_cambio'] is not None else 'N/D',
                data['ipc'] if data['ipc'] is not None else 'N/D',
            ]
            sheet.append(row_data)

    # Guardar el libro de trabajo en la respuesta HTTP
    workbook.save(response)
    return response

def export_economic_data_json(request):
    """
    Vista para exportar TODOS los datos económicos históricos a un archivo JSON.
    """
    # Obtener TODOS los hechos económicos históricos
    hechos = HechosEconomicos.objects.select_related('id_fecha', 'id_pais', 'id_indicador').order_by('id_fecha__anio', 'id_pais__nombre_pais', 'id_indicador__nombre_indicador')

    # Estructurar los datos en un formato de lista de diccionarios para JSON
    json_data_structured = []
    current_row = {}
    last_key = None # Para agrupar por Año, País, ISO

    for hecho in hechos:
        current_key = (hecho.id_fecha.anio, hecho.id_pais.nombre_pais, hecho.id_pais.codigo_iso)

        if current_key != last_key:
            if current_row: # Si ya hay una fila en construcción, añadirla a la lista
                json_data_structured.append(current_row)
            
            # Iniciar una nueva fila
            current_row = {
                'anio': hecho.id_fecha.anio,
                'pais': hecho.id_pais.nombre_pais,
                'codigo_iso': hecho.id_pais.codigo_iso,
                'inflacion': None,
                'crecimiento_pib': None,
                'tipo_cambio_usd_local': None,
                'ipc': None,
            }
            last_key = current_key
        
        # Asignar el valor del indicador a la clave correcta en la fila actual
        indicador_nombre = hecho.id_indicador.nombre_indicador
        if indicador_nombre == 'Inflación':
            current_row['inflacion'] = float(hecho.porcentaje_inflacion) if hecho.porcentaje_inflacion is not None else None
        elif indicador_nombre == 'Crecimiento PIB':
            current_row['crecimiento_pib'] = float(hecho.variacion_pib_anual) if hecho.variacion_pib_anual is not None else None
        elif indicador_nombre == 'Tipo de Cambio Dólar':
            current_row['tipo_cambio_usd_local'] = float(hecho.tipo_cambio_usd_local_promedio_cierre) if hecho.tipo_cambio_usd_local_promedio_cierre is not None else None
        elif indicador_nombre == 'IPC':
            current_row['ipc'] = float(hecho.ipc_o_devaluacion) if hecho.ipc_o_devaluacion is not None else None
    
    # Añadir la última fila si existe
    if current_row:
        json_data_structured.append(current_row)

    # Crear la respuesta HTTP con el JSON
    response = JsonResponse(json_data_structured, safe=False, json_dumps_params={'indent': 4})
    response['Content-Disposition'] = 'attachment; filename="datos_economicos_historicos.json"'
    return response

